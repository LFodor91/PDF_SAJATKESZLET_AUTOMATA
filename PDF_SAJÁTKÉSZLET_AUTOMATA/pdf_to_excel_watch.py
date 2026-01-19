import re
import time
from pathlib import Path

import pdfplumber
import openpyxl
from openpyxl.styles import PatternFill
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# =========================
# SETTINGS
# =========================
BASE_DIR = Path(__file__).resolve().parent

INPUT_DIR = BASE_DIR / "Input"
OUTPUT_DIR = BASE_DIR / "Output"

LOOKUP_FILE = BASE_DIR / "lookup.xlsx"
LOOKUP_SHEET_NAME = "Sheet1"  # ha más a lookup fül neve, írd át

# SABLON: a manuális mintád (zöld tábla téma)
TEMPLATE_FILE = BASE_DIR / "template.xlsx"
TEMPLATE_SHEET_NAME = "Sheet1"
TEMPLATE_TABLE_NAME = "Table1"

# lookup oszlopok (1-indexelt)
LOOKUP_COL_CODE = 2  # B (SAP-Cikkszám)
LOOKUP_COL_NAME = 4  # D (Megnevezés)

# fix output fájlnév (mindig felülír)
OUTPUT_FILENAME = "Sajatkeszlet.xlsx"

FILL_WARN = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


# =========================
# NORMALIZATION
# =========================
def normalize_code(value) -> str:
    """
    Lookup-ban lévő cikkszám normalizálása:
    - számblokkok összefűzése
    - amint elérünk egy 1-2 jegyű oszlopértékhez (pl. 1, 12), megállunk
    """
    if value is None:
        return ""
    tokens = re.findall(r"\d+", str(value))
    if not tokens:
        return ""
    out = []
    for t in tokens:
        if len(t) <= 2 and out:
            break
        out.append(t)
    return "".join(out)


def extract_code_key_from_line(after_pos: str) -> str:
    """
    Cikkszám kulcs a pozíció utáni részből:
    - elfogad >=6 számjegyű első blokkot
    - elfogad 5-3-4 mintát (14936 000 1000)
    - elfogad 5-3 mintát (09858 000 1)
    - megáll a "kis oszlop" értékeknél (1-2 jegy)
    """
    tokens = re.findall(r"\d+", after_pos)
    if not tokens:
        return ""

    ok = False
    if len(tokens[0]) >= 6:
        ok = True
    elif len(tokens) >= 3 and len(tokens[0]) == 5 and len(tokens[1]) == 3 and len(tokens[2]) == 4:
        ok = True
    elif len(tokens) >= 2 and len(tokens[0]) == 5 and len(tokens[1]) == 3:
        ok = True

    if not ok:
        return ""

    out = []
    for t in tokens:
        if len(t) <= 2 and out:
            break
        out.append(t)

    return "".join(out)




# =========================
# PDF
# =========================
def extract_text_from_pdf(pdf_path: Path) -> str:
    parts = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            parts.append(page.extract_text() or "")
    return "\n".join(parts)


def is_item_start_line(line: str) -> bool:
    """
    Tételkezdő sor:
    - elején pozíciószám (1-3 számjegy)
    - utána cikkszám:
        * >=6 számjegy
        * vagy 5-3-4 (14936 000 1000)
        * vagy 5-3 (09858 000 1)
    """
    m = re.match(r"^(\d{1,3})\s+(.*)$", line)
    if not m:
        return False

    after_pos = m.group(2).strip()
    tokens = re.findall(r"\d+", after_pos)
    if not tokens:
        return False

    if len(tokens[0]) >= 6:
        return True

    if len(tokens) >= 3 and len(tokens[0]) == 5 and len(tokens[1]) == 3 and len(tokens[2]) == 4:
        return True

    if len(tokens) >= 2 and len(tokens[0]) == 5 and len(tokens[1]) == 3:
        return True

    return False



def parse_pdf_items(pdf_text: str):
    """
    Robusztus tétel-parse (ugyanaz a logika, mint az EK programban):
    - tételkezdés: pozíció + >=6 számjegyű cikkszám
    - mennyiség:
        1) ugyanazon sorból: '(\d+) darab'
        2) ha nincs ott: előre nézünk max 20 sort,
           DE megállunk, ha új tétel kezdődik (nehogy rossz mennyiséget vegyünk fel)
    """
    lines = [ln.strip() for ln in pdf_text.splitlines() if ln.strip()]
    items = []

    for i, ln in enumerate(lines):
        low = ln.lower()

        # fejlécek / zajok
        if low.startswith("poz.") or "vevőszám" in low or low.startswith("megrendelés"):
            continue

        if not is_item_start_line(ln):
            continue

        after_pos = re.sub(r"^\d{1,3}\s+", "", ln).strip()
        key = extract_code_key_from_line(after_pos)
        if not key:
            continue

        qty = None

        # 1) ugyanazon sorból
        m = re.search(r"(\d+)\s+darab\b", low)
        if m:
            qty = int(m.group(1))
        else:
            # 2) előre nézés, de megállunk a következő tételnél
            for j in range(1, 21):
                if i + j >= len(lines):
                    break
                nxt = lines[i + j].strip()
                nxt_low = nxt.lower()

                if is_item_start_line(nxt):
                    break

                m2 = re.search(r"(\d+)\s+darab\b", nxt_low)
                if m2:
                    qty = int(m2.group(1))
                    break

        items.append({"key": key, "qty": qty})

    return items


# =========================
# LOOKUP
# =========================
def load_lookup_map():
    if not LOOKUP_FILE.exists():
        raise FileNotFoundError(f"Nem találom a lookup fájlt: {LOOKUP_FILE}")

    wb = openpyxl.load_workbook(LOOKUP_FILE, data_only=True)
    if LOOKUP_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Nincs ilyen sheet a lookupban: {LOOKUP_SHEET_NAME}. Van: {wb.sheetnames}")

    ws = wb[LOOKUP_SHEET_NAME]
    start_row = 3  # nálad így volt jó

    m = {}
    for r in range(start_row, ws.max_row + 1):
        sap = ws.cell(r, LOOKUP_COL_CODE).value
        name_hu = ws.cell(r, LOOKUP_COL_NAME).value

        k = normalize_code(sap)
        if not k:
            continue

        m[k] = {
            "code": str(sap).strip() if sap is not None else "",
            "name": str(name_hu).strip() if name_hu is not None else "",
        }

    return m


# =========================
# OUTPUT (TEMPLATE)
# =========================
def write_output_from_template(items, out_path: Path, lookup_map: dict):
    if not TEMPLATE_FILE.exists():
        raise FileNotFoundError(f"Nem találom a sablont: {TEMPLATE_FILE}")

    wb = openpyxl.load_workbook(TEMPLATE_FILE)
    if TEMPLATE_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Nincs ilyen sheet a sablonban: {TEMPLATE_SHEET_NAME}. Van: {wb.sheetnames}")

    ws = wb[TEMPLATE_SHEET_NAME]

    # régi adatok törlése (fejléc marad)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    out_row = 2
    missing = 0

    for it in items:
        hit = lookup_map.get(it["key"])
        qty = it.get("qty")

        if hit:
            ws.cell(out_row, 1).value = hit["code"]  # Cikkszám
            ws.cell(out_row, 2).value = hit["name"]  # Megnevezés
            ws.cell(out_row, 3).value = qty          # Mennyiség (PDF darab)
        else:
            missing += 1
            # Írjuk be a kulcsot, hogy lásd mi hiányzik a lookupból
            ws.cell(out_row, 1).value = it["key"]
            ws.cell(out_row, 3).value = qty
            for c in range(1, 4):
                ws.cell(out_row, c).fill = FILL_WARN

        out_row += 1

    # Table1 ref frissítés: 3 oszlop A:C
    last_row = max(1, ws.max_row)
    table_ref = f"A1:C{last_row}"

    table_obj = None
    # openpyxl verziótól függően
    if hasattr(ws, "tables"):
        for t in ws.tables.values():
            if getattr(t, "name", "") == TEMPLATE_TABLE_NAME or getattr(t, "displayName", "") == TEMPLATE_TABLE_NAME:
                table_obj = t
                break
    if table_obj is None and hasattr(ws, "_tables"):
        for t in ws._tables.values():
            if getattr(t, "name", "") == TEMPLATE_TABLE_NAME or getattr(t, "displayName", "") == TEMPLATE_TABLE_NAME:
                table_obj = t
                break

    if table_obj is None:
        raise ValueError(f"Nem találtam a táblát a sablonban: {TEMPLATE_TABLE_NAME}")

    table_obj.ref = table_ref
    if table_obj.autoFilter is not None:
        table_obj.autoFilter.ref = table_ref

    wb.save(out_path)
    return missing, len(items)


# =========================
# WATCHER
# =========================
class PdfDropHandler(FileSystemEventHandler):
    def __init__(self, lookup_map: dict):
        self.lookup_map = lookup_map

    def on_created(self, event):
        if event.is_directory:
            return

        p = Path(event.src_path)
        if p.suffix.lower() != ".pdf":
            return

        time.sleep(0.8)  # másolás befejezése

        try:
            pdf_text = extract_text_from_pdf(p)
            items = parse_pdf_items(pdf_text)

            out_path = OUTPUT_DIR / OUTPUT_FILENAME
            missing, total = write_output_from_template(items, out_path, self.lookup_map)

            print(f"[OK] {p.name} -> {out_path.name} | tételek: {total} | nem talált: {missing}")
        except Exception as e:
            print(f"[ERROR] {p.name}: {e}")


def main():
    INPUT_DIR.mkdir(exist_ok=True)
    OUTPUT_DIR.mkdir(exist_ok=True)

    lookup_map = load_lookup_map()

    observer = Observer()
    observer.schedule(PdfDropHandler(lookup_map), str(INPUT_DIR), recursive=False)
    observer.start()

    print(f"Watching: {INPUT_DIR}")
    print(f"Output:   {OUTPUT_DIR}")
    print("Drop a PDF into Input. Exit: Ctrl+C")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()


if __name__ == "__main__":
    main()
